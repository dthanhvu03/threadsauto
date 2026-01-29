"""
Module: content/excel_loader.py

Excel loader cho Threads automation.
Đọc nội dung từ Excel file và chuyển đổi thành jobs.
"""

# Standard library
from pathlib import Path
from typing import List, Dict, Any, Optional, TYPE_CHECKING
from datetime import datetime

# Third-party
if TYPE_CHECKING:
    import pandas as pd
else:
    try:
        import pandas as pd
    except ImportError:
        pd = None

# Local
from services.logger import StructuredLogger
from services.exceptions import ThreadsAutomationError
from utils.exception_utils import (
    safe_get_exception_type_name,
    safe_get_exception_message
)


class ExcelLoadError(ThreadsAutomationError):
    """Lỗi khi load Excel file."""


class ExcelLoader:
    """
    Loader cho Excel files.
    
    Format Excel file:
    - content (bắt buộc): Nội dung thread (tối đa 500 ký tự)
    - scheduled_time (tùy chọn): Thời gian lên lịch (format: YYYY-MM-DD HH:MM:SS hoặc YYYY-MM-DDTHH:MM:SS)
      **QUAN TRỌNG**: scheduled_time được coi là giờ Việt Nam (UTC+7), sẽ được tự động convert về UTC khi lưu
    - priority (tùy chọn): Độ ưu tiên (LOW, NORMAL, HIGH, URGENT)
    - platform (tùy chọn): Platform (THREADS, FACEBOOK), mặc định THREADS
    - link_aff (tùy chọn): Link affiliate sẽ được post riêng trong comment (chỉ cho THREADS)
    - cta (tùy chọn): Call-to-action sẽ được append vào cuối content
    - note (tùy chọn): Ghi chú nội bộ, không được sử dụng trong post
    """
    
    REQUIRED_COLUMNS = ["content"]
    OPTIONAL_COLUMNS = ["scheduled_time", "priority", "platform", "link_aff", "cta", "note"]
    VALID_PRIORITIES = ["LOW", "NORMAL", "HIGH", "URGENT"]
    VALID_PLATFORMS = ["THREADS", "FACEBOOK"]  # Tương ứng với Platform enum
    
    def __init__(self, logger: Optional[StructuredLogger] = None):
        """
        Khởi tạo Excel loader.
        
        Args:
            logger: Structured logger instance (tùy chọn)
        """
        if pd is None:
            raise ExcelLoadError(
                "pandas chưa được cài đặt. Chạy: pip install pandas openpyxl"
            )
        
        self.logger = logger or StructuredLogger(name="excel_loader")
    
    def load_from_file(self, file_path) -> List[Dict[str, Any]]:
        """
        Load dữ liệu từ Excel file.
        
        Args:
            file_path: Đường dẫn đến file Excel (.xlsx, .xls)
        
        Returns:
            List các dict với keys:
            - content (bắt buộc): Nội dung post
            - scheduled_time (optional): Thời gian lên lịch
            - priority (optional): Độ ưu tiên
            - platform (optional): Platform (THREADS, FACEBOOK)
            - link_aff (optional): Link affiliate (sẽ được post riêng trong comment)
            - cta (optional): Call-to-action (đã append vào content)
        
        Raises:
            ExcelLoadError: Nếu file không hợp lệ hoặc có lỗi khi đọc
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise ExcelLoadError(f"File không tồn tại: {file_path}")
        
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            raise ExcelLoadError(
                f"File không phải Excel format (.xlsx, .xls): {file_path.suffix}"
            )
        
        try:
            self.logger.log_step(
                step="LOAD_EXCEL",
                result="IN_PROGRESS",
                file_path=str(file_path)
            )
            
            # Đọc Excel file
            try:
                df = pd.read_excel(file_path, engine='openpyxl' if file_path.suffix == '.xlsx' else None)
            except PermissionError as e:
                raise ExcelLoadError(f"Không có quyền đọc file: {str(e)}") from e
            except (ValueError, OSError) as e:
                # ValueError: Invalid file format, corrupted file
                # OSError: File access issues (includes FileNotFoundError)
                raise ExcelLoadError(f"Lỗi đọc file Excel: {str(e)}") from e
            except Exception as e:
                # Catch any other pandas-specific errors (EmptyDataError, etc.)
                # Kiểm tra nếu là EmptyDataError
                error_msg = str(e)
                if "empty" in error_msg.lower() or "no data" in error_msg.lower():
                    raise ExcelLoadError("File Excel rỗng, không có dữ liệu") from e
                else:
                    raise ExcelLoadError(f"Lỗi đọc file Excel: {error_msg}") from e
            
            # Kiểm tra có dữ liệu không
            if df.empty:
                raise ExcelLoadError("File Excel rỗng, không có dữ liệu")
            
            # Kiểm tra cột bắt buộc
            missing_columns = [col for col in self.REQUIRED_COLUMNS if col not in df.columns]
            if missing_columns:
                raise ExcelLoadError(
                    f"Thiếu cột bắt buộc: {', '.join(missing_columns)}. "
                    f"Các cột hiện có: {', '.join(df.columns.tolist())}"
                )
            
            # Chuẩn hóa tên cột (loại bỏ khoảng trắng, chuyển lowercase)
            df.columns = df.columns.str.strip().str.lower()
            
            # Xử lý từng dòng
            posts = []
            skipped_rows = 0
            try:
                for index, row in df.iterrows():
                    try:
                        post_data = self._process_row(row)  # Process row
                        if post_data:
                            posts.append(post_data)
                    except Exception as e:
                        skipped_rows += 1
                        row_num = index + 2  # Excel row number (1-indexed + header)
                        
                        # Safe get error type name và message sử dụng utility functions
                        error_type_name = safe_get_exception_type_name(e)
                        error_msg = safe_get_exception_message(e, max_length=500)
                        
                        self.logger.log_step(
                            step="LOAD_EXCEL",
                            result="WARNING",
                            row=row_num,
                            error=f"Bỏ qua dòng {row_num}: {error_msg}",
                            error_type=error_type_name
                        )
                        continue
            except Exception as e:
                # Nếu iterrows() itself fails, wrap and re-raise
                raise ExcelLoadError(f"Lỗi khi đọc dữ liệu từ Excel: {str(e)}") from e
            
            if not posts:
                error_msg = (
                    f"Không có dòng dữ liệu hợp lệ nào trong file Excel. "
                    f"Tổng số dòng: {len(df)}, đã bỏ qua: {skipped_rows} dòng"
                )
                raise ExcelLoadError(error_msg)
            
            self.logger.log_step(
                step="LOAD_EXCEL",
                result="SUCCESS",
                file_path=str(file_path),
                total_rows=len(df),
                valid_posts=len(posts),
                skipped_rows=skipped_rows
            )
            
            return posts
            
        except ExcelLoadError:
            # Re-raise ExcelLoadError as-is
            raise
        except Exception as e:
            # Catch any other unexpected errors
            raise ExcelLoadError(f"Lỗi không mong đợi khi đọc Excel: {str(e)}") from e
    
    def _build_content(self, row: "pd.Series") -> str:
        """
        Xây dựng content từ các cột.
        
        Format: content + (cta nếu có)
        
        LƯU Ý: link_aff KHÔNG được append vào content nữa.
        link_aff sẽ được post riêng trong comment của bài viết.
        
        Args:
            row: Pandas Series chứa dữ liệu
        
        Returns:
            Content đã được build (không bao gồm link_aff)
        """
        content_parts = []
        
        # Content chính
        main_content = row.get("content", "")
        if not pd.isna(main_content) and str(main_content).strip() and str(main_content).strip().lower() != "nan":
            content_parts.append(str(main_content).strip())
        
        # CTA (nếu có) - vẫn append vào content
        cta = row.get("cta", "")
        if not pd.isna(cta) and str(cta).strip() and str(cta).strip().lower() not in ["nan", ""]:
            # Thêm CTA vào cuối content
            content_parts.append(str(cta).strip())
        
        # Join tất cả với newline
        final_content = "\n".join(content_parts)
        
        return final_content
    
    def _process_row(self, row: "pd.Series") -> Optional[Dict[str, Any]]:
        """
        Xử lý một dòng dữ liệu.
        
        Args:
            row: Pandas Series chứa dữ liệu của dòng
        
        Returns:
            Dict với post data (content, link_aff riêng, ...) hoặc None nếu dòng không hợp lệ
        """
        # Build content từ các cột (content + cta, KHÔNG bao gồm link_aff)
        content = self._build_content(row)
        
        # Validate content
        if not content or content.strip() == "":
            raise ValueError("Content không được để trống (sau khi build từ content/cta)")
        
        # Validate độ dài (Threads limit: 500 chars)
        # LƯU Ý: Validation chỉ áp dụng cho content chính (không bao gồm link_aff)
        if len(content) > 500:
            raise ValueError(
                f"Content quá dài ({len(content)} ký tự), tối đa 500 ký tự. "
                f"Preview: {content[:100]}..."
            )
        
        post_data = {
            "content": content
        }
        
        # Xử lý link_aff (tùy chọn) - TÁCH RIÊNG, không append vào content
        link_aff = row.get("link_aff", "")
        # Debug: Log để kiểm tra
        if link_aff is not None and not pd.isna(link_aff):
            link_aff_str = str(link_aff).strip()
            if link_aff_str and link_aff_str.lower() not in ["nan", ""]:
                post_data["link_aff"] = link_aff_str
                self.logger.log_step(
                    step="PROCESS_EXCEL_ROW",
                    result="INFO",
                    note=f"Found link_aff in row: {link_aff_str[:50]}..."
                )
        
        # Xử lý scheduled_time (tùy chọn)
        scheduled_time = row.get('scheduled_time', None)
        if not pd.isna(scheduled_time) and scheduled_time != '':
            try:
                # Hỗ trợ nhiều format
                if isinstance(scheduled_time, datetime):
                    post_data["scheduled_time"] = scheduled_time
                elif isinstance(scheduled_time, str):
                    scheduled_time_str = scheduled_time.strip()
                    # Thử các format
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"]:
                        try:
                            post_data["scheduled_time"] = datetime.strptime(scheduled_time_str, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        raise ValueError(f"Không thể parse scheduled_time: {scheduled_time_str}")
                else:
                    # Nếu là pandas Timestamp
                    post_data["scheduled_time"] = scheduled_time.to_pydatetime()
            except Exception as e:
                raise ValueError(f"Scheduled_time không hợp lệ: {str(e)}")
        
        # Xử lý priority (tùy chọn)
        priority = row.get('priority', None)
        if not pd.isna(priority) and priority != '':
            priority_str = str(priority).strip().upper()
            if priority_str not in self.VALID_PRIORITIES:
                raise ValueError(
                    f"Priority không hợp lệ: {priority_str}. "
                    f"Giá trị hợp lệ: {', '.join(self.VALID_PRIORITIES)}"
                )
            post_data["priority"] = priority_str
        
        # Xử lý platform (tùy chọn) - backward compatible: default THREADS nếu không có
        platform = row.get('platform', None)
        if not pd.isna(platform) and platform != '':
            platform_str = str(platform).strip().upper()
            if platform_str not in self.VALID_PLATFORMS:
                raise ValueError(
                    f"Platform không hợp lệ: {platform_str}. "
                    f"Giá trị hợp lệ: {', '.join(self.VALID_PLATFORMS)}"
                )
            post_data["platform"] = platform_str
        # Nếu không có platform, không thêm vào post_data - scheduler sẽ dùng default THREADS
        
        return post_data
    
    @staticmethod
    def create_template(output_path) -> None:
        """
        Tạo file Excel template mẫu đã chuẩn hóa.
        
        Template bao gồm:
        - Thứ tự cột logic: content -> scheduled_time -> priority -> platform -> link_aff -> cta -> note
        - Dữ liệu mẫu đầy đủ với các trường hợp sử dụng khác nhau
        - Format Excel với column width và header styling
        
        LƯU Ý VỀ link_aff:
        - link_aff sẽ được post riêng trong comment của bài viết (chỉ cho THREADS)
        - link_aff KHÔNG được append vào content
        - Nếu có link_aff, hệ thống sẽ tự động click "Thêm vào thread" và post link trong comment
        
        Args:
            output_path: Đường dẫn file Excel để tạo template
        """
        if pd is None:
            raise ExcelLoadError("pandas chưa được cài đặt")
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback nếu openpyxl không có styling support
            openpyxl_styling = False
        else:
            openpyxl_styling = True
        
        output_path = Path(output_path)
        
        # Tạo DataFrame mẫu với thứ tự cột chuẩn hóa
        # Thứ tự: content (bắt buộc) -> scheduled_time -> priority -> platform -> link_aff -> cta -> note
        template_data = {
            "content": [
                "Xin chào Threads! Đây là bài post đầu tiên sẽ được đăng ngay.",
                "Da khô bong tróc vào mùa hanh khô là nỗi ám ảnh của nhiều chị em. Vấn đề này không chỉ làm lớp nền không ăn phấn mà còn khiến da dễ bị kích ứng. Giải pháp: Hãy chú ý uống đủ nước, dùng máy tạo độ ẩm và chọn sữa rửa mặt có độ pH cân bằng. Đừng quên dưỡng ẩm ngay sau khi tắm nhé!",
                "Review sản phẩm skincare mới nhất - kem dưỡng ẩm có thành phần hyaluronic acid giúp da căng mọng, mềm mịn suốt cả ngày.",
                "Bài post được lên lịch với priority HIGH - sẽ được ưu tiên chạy trước.",
                "Bài post đầy đủ: có content chính + CTA + link affiliate (sẽ post trong comment riêng)",
                "Bài post cho Facebook platform - link_aff không được hỗ trợ (chỉ THREADS)"
            ],
            "scheduled_time": [
                "",  # Đăng ngay - để trống
                "2025-12-25 15:20:00",  # Lên lịch - format: YYYY-MM-DD HH:MM:SS
                "",  # Đăng ngay - để trống
                "2025-12-26 14:30:00",  # Lên lịch
                "",  # Đăng ngay - để trống
                ""   # Đăng ngay - để trống
            ],
            "priority": [
                "NORMAL",  # Mặc định (có thể để trống)
                "NORMAL",  # LOW, NORMAL, HIGH, URGENT
                "LOW",
                "HIGH",  # Priority cao - sẽ chạy trước các jobs khác cùng thời gian
                "NORMAL",
                "NORMAL"
            ],
            "platform": [
                "",  # Để trống = THREADS (mặc định)
                "",  # Để trống = THREADS (mặc định)
                "",  # Để trống = THREADS (mặc định)
                "",  # Để trống = THREADS (mặc định)
                "",  # Để trống = THREADS (mặc định)
                "FACEBOOK"  # Chỉ định platform FACEBOOK
            ],
            "link_aff": [
                "",  # Không có link affiliate
                "",  # Không có link affiliate
                "https://example.com/skincare?ref=aff123",  # Link affiliate - sẽ được post riêng trong comment
                "",  # Không có link affiliate
                "https://example.com/shop?ref=aff456",  # Link affiliate - sẽ được post riêng trong comment
                ""   # Không có link affiliate (FACEBOOK không hỗ trợ link_aff)
            ],
            "cta": [
                "",  # Không có CTA
                "Follow mình để xem thêm tips làm đẹp nhé ✨",  # CTA với emoji - sẽ được append vào cuối content
                "👉 Swipe để xem thêm review chi tiết!",  # CTA với emoji - sẽ được append vào cuối content
                "",  # Không có CTA
                "👉 Swipe để xem thêm tips! #skincare #beauty",  # CTA với hashtag - sẽ được append vào cuối content
                "Follow để cập nhật tin tức mới nhất!"  # CTA cho Facebook
            ],
            "note": [
                "Bài post đơn giản - chỉ có content, đăng ngay",  # Ghi chú
                "Làm đẹp, sức khỏe | Sharing (Prob-Sol) | Tip: chăm sóc da khô mùa đông | CTA sẽ append vào content",  # Ghi chú chi tiết
                "Workflow: 1) Post content chính 2) Click 'Thêm vào thread' 3) Post link_aff trong comment",  # Ghi chú workflow
                "Priority HIGH - sẽ được chạy trước các jobs khác cùng thời gian",  # Ghi chú
                "Ví dụ đầy đủ: content + cta (append vào post chính) + link_aff (post riêng trong comment)",  # Ghi chú
                "FACEBOOK platform - link_aff không được hỗ trợ (chỉ dành cho THREADS)"  # Ghi chú
            ]
        }
        
        # Tạo DataFrame với thứ tự cột chuẩn hóa
        df = pd.DataFrame(template_data)
        
        # Đảm bảo thứ tự cột: content -> scheduled_time -> priority -> platform -> link_aff -> cta -> note
        column_order = ["content", "scheduled_time", "priority", "platform", "link_aff", "cta", "note"]
        df = df[column_order]
        
        # Tạo file Excel
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Format Excel với openpyxl nếu có thể
        if openpyxl_styling:
            try:
                wb = load_workbook(output_path)
                ws = wb.active
                
                # Format header row
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for col_idx, col_name in enumerate(column_order, start=1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # Set column widths
                column_widths = {
                    "content": 60,  # Rộng nhất vì chứa nội dung dài
                    "scheduled_time": 20,
                    "priority": 12,
                    "platform": 12,
                    "link_aff": 40,
                    "cta": 35,
                    "note": 50
                }
                
                for col_idx, col_name in enumerate(column_order, start=1):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)
                
                # Wrap text cho cột content và note
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    # Content column (column 1)
                    content_cell = row[0]
                    content_cell.alignment = Alignment(wrap_text=True, vertical="top")
                    
                    # Note column (column 6)
                    note_cell = row[5]
                    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Set row height cho header
                ws.row_dimensions[1].height = 25
                
                # Freeze header row
                ws.freeze_panes = "A2"
                
                wb.save(output_path)
            except Exception:
                # Nếu formatting fail, vẫn giữ file Excel đã tạo
                pass

